import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import os

from utils.ui_helpers import render_header, render_empty_state, render_footer, branded_spinner
from database.database import get_dataset_metadata
from config.roles import Roles

from modules.export import _build_single_report_blocks, _render_excel_sheet
from utils.customer_metrics import money, percent

# --- Report Generators ---
def _gen_marketplace_summary(df):
    return _build_single_report_blocks(df, 100)

def _gen_sales_report(df):
    from utils.customer_metrics import build_sales_profile
    from modules.reports.sales_report import build_sales_report_blocks
    try:
        metrics = build_sales_profile(df, None)[1]
    except Exception:
        metrics = {}
    return build_sales_report_blocks(metrics)

def _gen_customer_report(df):
    from utils.customer_metrics import build_customer_profile
    from modules.reports.customer_report import build_customer_report_blocks
    try:
        metrics = build_customer_profile(df, None)[1]
    except Exception:
        metrics = {}
    return build_customer_report_blocks(metrics, None, None)

def _gen_inventory_report(df):
    from utils.customer_metrics import build_inventory_profile
    from modules.reports.inventory_report import build_inventory_report_blocks
    try:
        metrics = build_inventory_profile(df, None)[1]
    except Exception:
        metrics = {}
    return build_inventory_report_blocks(metrics)

def _gen_revenue_report(df):
    from modules.reports.payment_report import build_payment_report_blocks
    if 'gross_amount' in df.columns:
        revenue = df['gross_amount'].sum()
    elif 'total_amount' in df.columns:
        revenue = df['total_amount'].sum()
    elif 'line_revenue' in df.columns:
        revenue = df['line_revenue'].sum()
    elif 'payment_value' in df.columns:
        revenue = df['payment_value'].sum()
    elif 'price' in df.columns:
        revenue = df['price'].sum()
    else:
        revenue = 0
    metrics = {
        "total_revenue": revenue,
        "total_transactions": len(df),
        "success_rate": 1.0 - (len(df[df["payment_status"] == "Failed"]) / len(df) if len(df) > 0 and "payment_status" in df.columns else 0.0),
        "failed_count": 0
    }
    return build_payment_report_blocks(metrics)

def _gen_vendor_report(df):
    from utils.customer_metrics import build_vendor_profile
    from modules.reports.vendor_report import build_vendor_report_blocks
    try:
        profile, metrics, columns = build_vendor_profile(df, None)
        if not profile.empty:
            total_revenue = profile["revenue"].sum()
            if total_revenue > 0:
                profile["revenue_contribution_%"] = (profile["revenue"] / total_revenue * 100).round(2)
            else:
                profile["revenue_contribution_%"] = 0.0

            top_vendors = profile.sort_values(by="revenue", ascending=False).head(10)
            
            display_cols = ["vendor", "revenue", "revenue_contribution_%", "orders", "avg_order_value"]
            if "avg_rating" in profile.columns and not profile["avg_rating"].isna().all():
                display_cols.append("avg_rating")
            
            top_vendors = top_vendors[[c for c in display_cols if c in top_vendors.columns]]
            
            top_vendors["revenue"] = top_vendors["revenue"].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "$0.00")
            if "avg_order_value" in top_vendors:
                top_vendors["avg_order_value"] = top_vendors["avg_order_value"].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "$0.00")
                
            metrics["top_vendor_revenue"] = profile["revenue"].max()
            metrics["lowest_vendor_revenue"] = profile["revenue"].min()
        else:
            top_vendors = pd.DataFrame()
            
    except Exception:
        metrics = {}
        top_vendors = pd.DataFrame()
        
    return build_vendor_report_blocks(metrics, top_vendors=top_vendors)

def _gen_order_report(df):
    from modules.reports.order_report import build_order_report_blocks
    total = len(df)
    rate = 0.05
    return build_order_report_blocks(total, rate)

def _gen_pnl_report(df):
    from modules.reports.report_utils import build_module_report
    from utils.customer_metrics import money
    if 'gross_amount' in df.columns:
        revenue = df['gross_amount'].sum()
    elif 'total_amount' in df.columns:
        revenue = df['total_amount'].sum()
    elif 'line_revenue' in df.columns:
        revenue = df['line_revenue'].sum()
    elif 'payment_value' in df.columns:
        revenue = df['payment_value'].sum()
    elif 'price' in df.columns:
        revenue = df['price'].sum()
    else:
        revenue = 0
    has_cost = 'cost' in df.columns or 'cogs' in df.columns
    has_expense = 'marketing_expense' in df.columns or 'operations_expense' in df.columns
    
    metrics = [("Gross Revenue", money(revenue))]
    insights = []
    
    if has_cost and has_expense:
        cost = df['cost'].sum() if 'cost' in df.columns else df['cogs'].sum()
        expense = (df['marketing_expense'].sum() if 'marketing_expense' in df.columns else 0) + \
                  (df['operations_expense'].sum() if 'operations_expense' in df.columns else 0)
        metrics.append(("Net Profit", money(revenue - cost - expense)))
    else:
        metrics.append(("Net Profit", "N/A"))
        insights.append("Financial metric unavailable due to insufficient source data. Missing cost or expense fields.")
        
    return build_module_report("Profit & Loss Report", metrics=metrics, insights=insights)

def _gen_expense_report(df):
    from modules.reports.report_utils import build_module_report
    from utils.customer_metrics import money
    has_expense = 'marketing_expense' in df.columns or 'operations_expense' in df.columns
    
    metrics = []
    insights = []
    
    if has_expense:
        metrics.append(("Marketing Exp.", money(df['marketing_expense'].sum() if 'marketing_expense' in df.columns else 0)))
        metrics.append(("Ops Exp.", money(df['operations_expense'].sum() if 'operations_expense' in df.columns else 0)))
    else:
        metrics.append(("Marketing Exp.", "N/A"))
        metrics.append(("Ops Exp.", "N/A"))
        insights.append("Financial metric unavailable due to insufficient source data. Missing expense fields.")
        
    return build_module_report("Expense Analysis Report", metrics=metrics, insights=insights)

def _gen_refund_report(df):
    from modules.reports.report_utils import build_module_report
    from utils.customer_metrics import percent, money
    if 'gross_amount' in df.columns:
        revenue = df['gross_amount'].sum()
    elif 'total_amount' in df.columns:
        revenue = df['total_amount'].sum()
    elif 'line_revenue' in df.columns:
        revenue = df['line_revenue'].sum()
    elif 'payment_value' in df.columns:
        revenue = df['payment_value'].sum()
    elif 'price' in df.columns:
        revenue = df['price'].sum()
    else:
        revenue = 0
    has_refund = 'refund_amount' in df.columns or 'is_refunded' in df.columns
    
    metrics = []
    insights = []
    
    if has_refund:
        refund_amt = df['refund_amount'].sum() if 'refund_amount' in df.columns else 0
        refund_rate = refund_amt / revenue if revenue > 0 else 0
        metrics.append(("Refund Rate", percent(refund_rate)))
        metrics.append(("Refund Amount", money(refund_amt)))
    else:
        metrics.append(("Refund Rate", "N/A"))
        metrics.append(("Refund Amount", "N/A"))
        insights.append("Financial metric unavailable due to insufficient source data. Missing refund fields.")
        
    return build_module_report("Refund Analysis Report", metrics=metrics, insights=insights)


# --- Context Mappings ---
REPORT_CONTEXTS = {
    "Marketplace Reports": {
        "title": "Marketplace Reports",
        "subtitle": "Generate, filter, and export overall marketplace reports.",
        "reports": {
            "Executive summary": {"desc": "High-level overview of marketplace health.", "metrics": "Revenue, Orders, Vendors", "audience": "C-Level, Admins", "dataset": "marketplace", "generator": _gen_marketplace_summary},
            "Sales": {"desc": "Detailed analysis of sales trends.", "metrics": "Sales Volume, AOV, Conversion", "audience": "Sales Managers", "dataset": "sales", "generator": _gen_sales_report},
            "Customers": {"desc": "Customer growth and retention metrics.", "metrics": "Active Customers, Churn", "audience": "Marketing", "dataset": "customer", "generator": _gen_customer_report},
            "Inventory": {"desc": "Marketplace inventory health.", "metrics": "Stock Levels, Out of Stock", "audience": "Supply Chain", "dataset": "inventory", "generator": _gen_inventory_report},
            "Revenue": {"desc": "Top-line revenue and profitability.", "metrics": "Gross Revenue, Net Margin", "audience": "Finance", "dataset": "payment", "generator": _gen_revenue_report}
        },
        "filters": ["date_range", "vendor", "category", "region"]
    },
    "Admin Vendor Reports": {
        "title": "Vendor Reports",
        "subtitle": "Generate, filter, and export vendor-specific reports.",
        "reports": {
            "Vendor performance": {"desc": "Overall performance ranking of vendors.", "metrics": "Score, Fulfillment Rate", "audience": "Vendor Managers", "dataset": "vendor", "generator": _gen_vendor_report},
            "Vendor sales": {"desc": "Sales broken down by vendor.", "metrics": "Sales Volume by Vendor", "audience": "Sales Managers", "dataset": "sales", "generator": _gen_sales_report},
            "Vendor revenue": {"desc": "Revenue generation by vendor.", "metrics": "Revenue, Take Rate", "audience": "Finance", "dataset": "payment", "generator": _gen_revenue_report},
            "Vendor orders": {"desc": "Order volume and status per vendor.", "metrics": "Order Count, Status", "audience": "Operations", "dataset": "order", "generator": _gen_order_report},
            "Top vendors": {"desc": "List of the highest-performing vendors.", "metrics": "Top Revenue, Top Volume", "audience": "C-Level, Admins", "dataset": "vendor", "generator": _gen_vendor_report}
        },
        "filters": ["date_range", "vendor", "product", "category"]
    },
    "Financial Reports": {
        "title": "Financial Reports",
        "subtitle": "Generate, filter, and export financial reporting.",
        "reports": {
            "Profit & Loss": {"desc": "Comprehensive P&L statement.", "metrics": "Gross Profit, Net Profit", "audience": "Finance, C-Level", "dataset": "payment", "generator": _gen_pnl_report},
            "Revenue summary": {"desc": "Summarized revenue streams.", "metrics": "Revenue by Category", "audience": "Finance", "dataset": "payment", "generator": _gen_revenue_report},
            "Expense analysis": {"desc": "Breakdown of marketplace expenses.", "metrics": "Marketing, Operations costs", "audience": "Finance", "dataset": "payment", "generator": _gen_expense_report},
            "Payment reports": {"desc": "Status of customer payments.", "metrics": "Success Rate, Methods", "audience": "Accounting", "dataset": "payment", "generator": _gen_revenue_report},
            "Refund reports": {"desc": "Analysis of refunded orders.", "metrics": "Refund Rate, Amount", "audience": "Accounting, Support", "dataset": "payment", "generator": _gen_refund_report}
        },
        "filters": ["date_range", "payment_status", "payment_method", "refund_status"]
    },
    "Reports": {
        "title": "My Store Reports",
        "subtitle": "Generate, filter, and export store-specific performance reports.",
        "reports": {
            "Sales Report": {"desc": "Your overall sales performance.", "metrics": "Sales Volume, AOV", "audience": "Store Owner", "dataset": "sales", "generator": _gen_sales_report},
            "Revenue Report": {"desc": "Your revenue and payouts.", "metrics": "Net Revenue, Fees", "audience": "Store Owner", "dataset": "payment", "generator": _gen_revenue_report},
            "Inventory Report": {"desc": "Your current inventory health.", "metrics": "Stock Levels", "audience": "Store Manager", "dataset": "inventory", "generator": _gen_inventory_report},
            "Product Performance Report": {"desc": "Performance of your individual products.", "metrics": "Product Sales", "audience": "Store Manager", "dataset": "sales", "generator": _gen_sales_report},
            "Settlement Report": {"desc": "Status of your payouts from the marketplace.", "metrics": "Pending Payouts, Settled", "audience": "Store Owner", "dataset": "payment", "generator": _gen_revenue_report}
        },
        "filters": ["date_range", "product", "category"]
    }
}

def _generate_excel_bytes(blocks, sheet_title="Report"):
    import openpyxl
    excel_buffer = BytesIO()
    workbook = openpyxl.Workbook()
    
    if "Sheet" in workbook.sheetnames:
        worksheet = workbook["Sheet"]
        worksheet.title = sheet_title
    else:
        worksheet = workbook.create_sheet(title=sheet_title)
        
    _render_excel_sheet(worksheet, blocks)
    
    worksheet.column_dimensions['A'].width = 90
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 30
    
    workbook.save(excel_buffer)
    return excel_buffer.getvalue()

def show_report_center():
    nav = st.session_state.get('current_nav', 'Reports')
    role = st.session_state.get('role', Roles.MANAGER)
    
    if nav not in REPORT_CONTEXTS:
        if role == Roles.VENDOR:
            nav = "Reports"
        else:
            nav = "Marketplace Reports"
            
    context = REPORT_CONTEXTS[nav]
    
    render_header(context["title"], context["subtitle"], "Reports")
    
    st.markdown("### Report Configuration")
    report_names = list(context["reports"].keys())
    
    col1, col2 = st.columns([2, 1])
    with col1:
        selected_report = st.selectbox("Select Report Type", report_names)
    with col2:
        export_format = st.selectbox("Export Format", ["Excel (.xlsx)", "CSV (.csv)", "PDF (.pdf)"])
        
    report_meta = context["reports"][selected_report]
    
    # Context-aware data fetching
    from utils.data_source_helper import get_analytics_df, render_data_source_banner
    df, source_label, source_name = get_analytics_df(report_meta["dataset"])
    render_data_source_banner(source_label, source_name)
    
    if df is None or df.empty:
        st.warning("Data not available for this report type.")
        render_empty_state()
        render_footer("Reports")
        return
        
    # --- Report Information Panel ---
    try:
        cont_info = st.container(border=True)
    except TypeError:
        cont_info = st.container()
        
    with cont_info:
        st.markdown(f"**Description:** {report_meta['desc']}")
        c1, c2 = st.columns(2)
        c1.markdown(f"**Key Metrics:** {report_meta['metrics']}")
        c2.markdown(f"**Target Audience:** {report_meta['audience']}")
        
    # --- Dynamic Filters ---
    st.markdown("### Filters")
    active_filters = {}
    
    with st.expander("Apply Data Filters", expanded=True):
        f_cols = st.columns(min(len(context["filters"]), 4) or 1)
        
        for idx, f_type in enumerate(context["filters"]):
            col = f_cols[idx % len(f_cols)]
            with col:
                if f_type == "date_range":
                    date_range = st.date_input("Date Range", [])
                    if date_range:
                        active_filters["Date Range"] = f"{date_range[0]} to {date_range[-1]}"
                elif f_type == "vendor":
                    if role != Roles.VENDOR:
                        opts = ["All Vendors"] + (df['seller_id'].dropna().unique().tolist() if 'seller_id' in df.columns else [])
                        val = st.selectbox("Vendor Filter", opts)
                        if val != "All Vendors": active_filters["Vendor"] = val
                    else:
                        st.text_input("Vendor", value="Your Store", disabled=True)
                        active_filters["Vendor"] = "Your Store"
                elif f_type == "category":
                    opts = ["All Categories"] + (df['product_category_name'].dropna().unique().tolist() if 'product_category_name' in df.columns else [])
                    val = st.selectbox("Category Filter", opts)
                    if val != "All Categories": active_filters["Category"] = val
                elif f_type == "region":
                    opts = ["All Regions"] + (df['customer_state'].dropna().unique().tolist() if 'customer_state' in df.columns else [])
                    val = st.selectbox("Region", opts)
                    if val != "All Regions": active_filters["Region"] = val
                elif f_type == "product":
                    opts = ["All Products"] + (df['product_id'].dropna().unique().tolist() if 'product_id' in df.columns else [])
                    val = st.selectbox("Product", opts)
                    if val != "All Products": active_filters["Product"] = val
                elif f_type == "payment_status":
                    val = st.selectbox("Payment Status", ["All", "Paid", "Pending", "Failed"])
                    if val != "All": active_filters["Payment Status"] = val
                elif f_type == "payment_method":
                    opts = ["All Methods"] + (df['payment_type'].dropna().unique().tolist() if 'payment_type' in df.columns else [])
                    val = st.selectbox("Payment Method", opts)
                    if val != "All Methods": active_filters["Payment Method"] = val
                elif f_type == "refund_status":
                    val = st.selectbox("Refund Status", ["All", "Refunded", "Not Refunded"])
                    if val != "All": active_filters["Refund Status"] = val

    # --- Apply filtering to df ---
    filtered_df = df.copy()
    if "Vendor" in active_filters and active_filters["Vendor"] != "Your Store" and "seller_id" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["seller_id"] == active_filters["Vendor"]]
    if "Category" in active_filters and "product_category_name" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["product_category_name"] == active_filters["Category"]]
    if "Region" in active_filters and "customer_state" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["customer_state"] == active_filters["Region"]]
        
    record_count = len(filtered_df)

    # --- Report Summary & Validation Panel ---
    st.markdown("### Report Status")
    try:
        cont_summary = st.container(border=True)
    except TypeError:
        cont_summary = st.container()
        
    with cont_summary:
        s1, s2, s3, s4 = st.columns(4)
        s1.markdown(f"**Type:**<br>{selected_report}", unsafe_allow_html=True)
        s2.markdown(f"**Data Source:**<br>{source_name}", unsafe_allow_html=True)
        s3.markdown(f"**Format:**<br>{export_format}", unsafe_allow_html=True)
        
        filter_str = ", ".join([f"{k}: {v}" for k, v in active_filters.items()]) if active_filters else "None"
        s4.markdown(f"**Filters:**<br>{filter_str}", unsafe_allow_html=True)
        
        st.markdown("---")
        if record_count > 0:
            st.success(f"**Ready for Export:** {record_count:,} records found matching your filters.")
        else:
            st.error("**No Data Available:** 0 records found. Please adjust your filters.")
            
    # --- Generate & Export ---
    generate_disabled = record_count == 0
    
    if st.button(f"Generate & Export {selected_report}", type="primary", use_container_width=True, disabled=generate_disabled):
        with branded_spinner("Compiling Professional Report..."):
            
            meta_df = pd.DataFrame({
                "Property": ["Report Name", "Generated On", "Generated By", "Applied Filters", "Total Records Exported", "Data Source"],
                "Value": [
                    selected_report,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    st.session_state.get("full_name", role),
                    filter_str,
                    str(record_count),
                    source_name
                ]
            })
            
            blocks = []
            blocks.append({"type": "title", "text": f"{context['title']} - {selected_report}"})
            blocks.append({"type": "h1", "text": "Report Metadata"})
            blocks.append({"type": "table", "title": "Report Details", "df": meta_df})
            
            # Dynamic block generation
            generator_func = report_meta["generator"]
            try:
                standard_blocks = generator_func(filtered_df)
                if standard_blocks and standard_blocks[0]["type"] == "title":
                    standard_blocks = standard_blocks[1:]
                blocks.extend(standard_blocks)
            except Exception as e:
                import traceback
                traceback.print_exc()
                st.error("Failed to generate report content due to missing metric calculation data.")
                return
            
            context_prefix = nav.replace(" ", "_").lower()
            report_slug = selected_report.replace(" ", "_").lower()
            today_str = datetime.date.today().strftime("%Y-%m-%d")
            
            if export_format == "Excel (.xlsx)":
                file_bytes = _generate_excel_bytes(blocks, sheet_title="Report")
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ext = "xlsx"
                filename = f"{context_prefix}_{report_slug}_{today_str}.xlsx"
            else:
                file_bytes = filtered_df.to_csv(index=False).encode('utf-8')
                mime = "text/csv"
                ext = "csv"
                filename = f"{context_prefix}_{report_slug}_{today_str}.csv"
                
        st.success(f"Report generated successfully!")
        
        st.download_button(
            label=f"Download {filename}",
            data=file_bytes,
            file_name=filename,
            mime=mime,
            use_container_width=True,
            type="secondary"
        )
        
    render_footer("Reports")
