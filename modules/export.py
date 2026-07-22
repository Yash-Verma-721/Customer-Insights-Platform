import time
import streamlit as st
import pandas as pd
from io import BytesIO
import os
import datetime

from utils.customer_metrics import detect_customer_columns, money, percent
from utils.ui_helpers import (
    render_header, render_empty_state, render_help_expander, 
    render_footer, add_session_log, show_smart_notification,
    calculate_readiness_score, branded_spinner
)

def _build_single_report_blocks(df, score):
    blocks = []
    
    from utils.cache import get_cached_metric
    from utils.customer_metrics import (
        detect_customer_columns, detect_marketplace_columns,
        build_customer_profile, build_vendor_profile, build_inventory_profile, build_sales_profile,
        build_recommendations
    )
    from modules.analytics_order import _process_order_data
    from modules.analytics_payment import _process_payment_data
    from utils.ml_models import get_ml_customer_metrics, get_ml_customer_summary, get_ml_customer_recommendations, ml_sales_forecast
    
    detected_cust = get_cached_metric("detected", detect_customer_columns, df)
    detected_mp = get_cached_metric("detected_mp", detect_marketplace_columns, df)
    
    cust_profile, cust_metrics, columns = get_cached_metric("profile", build_customer_profile, df, detected_cust)
    vendor_profile, vendor_metrics, _ = get_cached_metric("vendor_profile", build_vendor_profile, df, detected_mp)
    inv_profile, inv_metrics, _ = get_cached_metric("inv_profile", build_inventory_profile, df, detected_mp)
    sales_profile, sales_metrics, _ = get_cached_metric("sales_profile", build_sales_profile, df, detected_mp)
    total_orders, order_status_counts, order_status_df, _, monthly_orders, has_status, has_date = get_cached_metric("order_profile", _process_order_data, df, detected_mp)
    payment_metrics = get_cached_metric("payment_profile", _process_payment_data, df, detected_mp)
    
    ml_metrics = get_ml_customer_metrics()
    ml_summ = get_ml_customer_summary()
    recs, _ = get_cached_metric("dashboard_recs", build_recommendations, df, detected_mp)
    
    forecast = None
    if sales_metrics.get("has_date_column") and not sales_profile.empty and "date" in sales_profile.columns:
        profile_dates = pd.to_datetime(sales_profile["date"], errors="coerce")
        profile_dates_period = profile_dates.dt.to_period("M").astype(str)
        monthly_summary = sales_profile.groupby(profile_dates_period).agg(Revenue=("revenue", "sum")).reset_index().rename(columns={"date": "Month"})
        forecast = ml_sales_forecast(monthly_summary)

    section_num = 1
    
    blocks.append({"type": "title", "text": "Customer Insights Report"})
    
    # 1. Executive Summary
    blocks.append({"type": "h1", "text": f"{section_num}. Executive Summary"})
    blocks.append({"type": "table", "title": "Executive KPIs", "df": pd.DataFrame({
        "Metric": ["Dataset Health", "Total Revenue", "Total Orders", "Customers", "Vendors", "Products", "Inventory Value"],
        "Value": [f"{score}/100", money(sales_metrics.get("total_sales", 0)), f"{total_orders:,}", f"{cust_metrics.get('total_customers', 0):,}", f"{vendor_metrics.get('total_vendors', 0):,}", f"{inv_metrics.get('total_products', 0):,}", money(inv_metrics.get("inventory_value", 0))]
    })})
    section_num += 1
    
    # 2. Sales Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Sales Analytics"})
    blocks.append({"type": "table", "title": "Sales Metrics", "df": pd.DataFrame({
        "Metric": ["Avg Order Value", "Growth %"],
        "Value": [money(sales_metrics.get("avg_order_value", 0)), percent(sales_metrics.get("growth_pct", 0))]
    })})
    section_num += 1
    
    # 3. Customer Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Customer Analytics"})
    blocks.append({"type": "table", "title": "Customer Metrics", "df": pd.DataFrame({
        "Metric": ["Repeat Purchase Rate", "Avg Frequency"],
        "Value": [percent(cust_metrics.get("repeat_rate", 0)), f"{cust_metrics.get('purchase_frequency', 1.0):.2f}x"]
    })})
    section_num += 1

    # 4. Vendor Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Vendor Analytics"})
    blocks.append({"type": "table", "title": "Vendor Metrics", "df": pd.DataFrame({
        "Metric": ["Total Vendors", "Top Vendor"],
        "Value": [f"{vendor_metrics.get('total_vendors', 0):,}", str(vendor_metrics.get('top_vendor_by_revenue', 'N/A'))]
    })})
    section_num += 1
    
    # 5. Product Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Product Analytics"})
    blocks.append({"type": "table", "title": "Product Metrics", "df": pd.DataFrame({
        "Metric": ["Total Products", "Categories", "Avg Price"],
        "Value": [f"{inv_metrics.get('total_products', 0):,}", f"{inv_metrics.get('total_categories', 0):,}", money(inv_metrics.get('avg_price', 0))]
    })})
    section_num += 1
    
    # 6. Inventory Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Inventory Analytics"})
    blocks.append({"type": "table", "title": "Inventory Metrics", "df": pd.DataFrame({
        "Metric": ["Total Value", "Low Stock Items", "Out of Stock"],
        "Value": [money(inv_metrics.get('inventory_value', 0)), f"{inv_metrics.get('low_stock_count', 0):,}", f"{inv_metrics.get('out_of_stock_count', 0):,}"]
    })})
    section_num += 1
    
    # 7. Order Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Order Analytics"})
    if has_status and not order_status_df.empty:
        blocks.append({"type": "table", "title": "Order Fulfillment", "df": order_status_df})
    else:
        blocks.append({"type": "p", "text": "No status data available."})
    section_num += 1
    
    # 8. Payment Analytics
    blocks.append({"type": "h1", "text": f"{section_num}. Payment Analytics"})
    blocks.append({"type": "table", "title": "Payment Metrics", "df": pd.DataFrame({
        "Metric": ["Total Methods", "Success Rate", "Failed Payments"],
        "Value": [f"{payment_metrics.get('total_methods', 0)}", percent(payment_metrics.get('success_rate', 0)), f"{payment_metrics.get('failed_count', 0):,}"]
    })})
    section_num += 1
    
    # 9. ML Segmentation
    blocks.append({"type": "h1", "text": f"{section_num}. ML Segmentation"})
    if ml_metrics and ml_metrics.get("status") == "success" and ml_summ is not None and not ml_summ.empty:
        blocks.append({"type": "table", "title": "Customer Segments", "df": ml_summ[["ml_segment", "customer_count", "percentage"]]})
    else:
        blocks.append({"type": "p", "text": "ML Segmentation not executed or unavailable."})
    section_num += 1
    
    # 10. Sales Forecast
    blocks.append({"type": "h1", "text": f"{section_num}. Sales Forecast"})
    if forecast and forecast.get("status") == "success":
        blocks.append({"type": "table", "title": "AI Sales Forecast", "df": pd.DataFrame({
            "Metric": ["Next Period Revenue", "Trend", "Confidence", "Recommendation"],
            "Value": [money(forecast["forecast_revenue"]), forecast["trend"], forecast["confidence"], forecast["recommendation"]]
        })})
    else:
        blocks.append({"type": "p", "text": "Forecast not available."})
    section_num += 1
    
    # 11. AI Recommendations
    blocks.append({"type": "h1", "text": f"{section_num}. AI Recommendations"})
    if recs:
        items = [f"[{r.get('priority', 'Low')}] {r.get('title', 'Insight')} - {r.get('message', '')}" for r in recs[:5]]
        blocks.append({"type": "bullets", "items": items})
    else:
        blocks.append({"type": "p", "text": "No recommendations available."})
        
    return blocks

def _render_excel_sheet(worksheet, blocks):
    from openpyxl.styles import Font, Alignment, PatternFill
    
    current_row = 1
    for b in blocks:
        if b["type"] == "title":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=20, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            worksheet.row_dimensions[current_row].height = 35
            current_row += 3
        elif b["type"] == "h1":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=16, bold=True, color="1E293B")
            current_row += 2
        elif b["type"] == "h2":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=12, bold=True, color="334155")
            current_row += 1
        elif b["type"] == "p":
            lines = str(b["text"]).split('\n')
            for line in lines:
                cell = worksheet.cell(row=current_row, column=1, value=line.strip())
                cell.font = Font(size=11, color="475569")
                current_row += 1
            current_row += 1
        elif b["type"] == "bullets":
            for item in b["items"]:
                cell = worksheet.cell(row=current_row, column=1, value=f"- {item}")
                cell.font = Font(size=11, color="475569")
                current_row += 1
            current_row += 1
        elif b["type"] == "table":
            cell = worksheet.cell(row=current_row, column=1, value=b["title"])
            cell.font = Font(bold=True, size=11, color="0F172A")
            current_row += 1
            
            df = b["df"]
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=str(col_name))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            current_row += 1
            
            for _, row in df.iterrows():
                for col_idx, val in enumerate(row, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=str(val))
                    cell.font = Font(size=11, color="334155")
                current_row += 1
            current_row += 2

def show_export():
    render_header(
        "Customer Insights Report Generator",
        "Generate and download a professional Management Report.",
        "Export Reports"
    )
    
    render_help_expander(
        "Generate a comprehensive, single-document Management Report. "
        "The report includes an Executive Summary, Key Findings, Interpretation, and Strategic Recommendations across all modules. "
        "It is designed to be immediately readable by stakeholders without navigating complex spreadsheets."
    )

    if "df" not in st.session_state or st.session_state["df"] is None or st.session_state["df"].empty:
        st.warning("Please upload a dataset before generating the management report.")
        render_empty_state()
        render_footer("Export Reports")
        return

    df = st.session_state["df"]
    score = calculate_readiness_score(df)
    role = st.session_state.get('role', 'Manager')
    
    st.markdown("## Management Report")
    
    from database.database import get_dataset_metadata, update_report_metadata

    user_id = st.session_state.get("user_id")
    if user_id:
        user_dir = f"datasets/user_{user_id}"
        os.makedirs(user_dir, exist_ok=True)
        report_path = f"{user_dir}/customer_insights_report.xlsx"
        meta = get_dataset_metadata(user_id=user_id)
    else:
        report_path = "datasets/active_report.xlsx"
        meta = get_dataset_metadata()
    
    st.markdown("### Report Status")
    try:
        cont_status = st.container(border=True)
    except TypeError:
        cont_status = st.container()
        
    with cont_status:
        if meta and meta.get("last_report_time"):
            r1, r2, r3, r4, r5 = st.columns(5)
            report_filename = os.path.basename(report_path)
            with r1: st.markdown(f"**Latest Report:**<br>{report_filename}", unsafe_allow_html=True)
            with r2: st.markdown(f"**Generated By:**<br>{meta.get('last_report_by')}", unsafe_allow_html=True)
            with r3: st.markdown(f"**Generated On:**<br>{meta.get('last_report_time')}", unsafe_allow_html=True)
            with r4: st.markdown(f"**Dataset Used:**<br>{meta.get('dataset_name')}", unsafe_allow_html=True)
            with r5: st.markdown(f"**Status:**<br>Published", unsafe_allow_html=True)
        else:
            st.info("No report has been published yet.")

    if role == "Manager":
        st.markdown("### Download Reports")
        try:
            cont_dl = st.container(border=True)
        except TypeError:
            cont_dl = st.container()
            
        with cont_dl:
            report_filename = os.path.basename(report_path)
            if os.path.exists(report_path):
                st.success("The latest Management Report is available for download.")
                with open(report_path, "rb") as f:
                    report_data = f.read()
                
                st.download_button(
                    label="Download Customer Insights Report",
                    data=report_data,
                    file_name=report_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            else:
                st.info("No report has been published yet. Please contact your Analyst.")
        render_footer("Export Reports")
        return

    st.markdown("### Report Configuration")
    try:
        cont_config = st.container(border=True)
    except TypeError:
        cont_config = st.container()
        
    with cont_config:
        st.write("Export a consolidated analytics narrative compiled into a single, professional Excel document.")
        st.info("The report systematically includes Executive Summaries and module-level analyses (Sales, Customer, Vendor, Product, Inventory, Orders, Payments) plus AI-driven recommendations.")
    
        is_cleaned = st.session_state.get("is_cleaned", False)
        if not is_cleaned:
            st.warning("Complete Data Cleaning before generating and publishing the management report.")
            st.button("Generate & Publish Management Report", type="primary", disabled=True)
        else:
            if st.button("Generate & Publish Management Report", type="primary"):
                with branded_spinner("Compiling Analytics Data..."):
                    blocks = _build_single_report_blocks(df, score)
                    
                excel_buffer = BytesIO()
                try:
                    with branded_spinner("Generating Excel Report..."):
                        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                            workbook = writer.book
                            worksheet = workbook.create_sheet(title="Platform Report")
                            
                            _render_excel_sheet(worksheet, blocks)
                            
                            if "Sheet" in workbook.sheetnames:
                                del workbook["Sheet"]
                            
                            worksheet.column_dimensions['A'].width = 90
                            worksheet.column_dimensions['B'].width = 30
                            worksheet.column_dimensions['C'].width = 30
                            
                        excel_data = excel_buffer.getvalue()
                        
                        os.makedirs(os.path.dirname(report_path), exist_ok=True)
                        with open(report_path, "wb") as f:
                            f.write(excel_data)
                        
                        gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        gen_by = st.session_state.get("full_name", "Analyst")
                        update_report_metadata(gen_time, gen_by, user_id=user_id)
                        
                        from database.database import publish_dataset
                        df_to_publish = st.session_state.get("df")
                        publish_dataset(user_id, df_to_publish)
                        
                        if "file_details" in st.session_state:
                            st.session_state["file_details"]["status"] = "Published"
                        
                        add_session_log(f"Generated & Published Module-Wide Management Report")
                        st.success("Management Report generated successfully.\n\n- All modules exported.\n- Cleaned dataset published.\n- Managers can now access the latest approved dataset and report.")
                        st.rerun()
                        
                except Exception as e:
                    show_smart_notification("error", f"Report generation failed: {e}")

    st.markdown("### Personal Download")
    try:
        cont_dl_analyst = st.container(border=True)
    except TypeError:
        cont_dl_analyst = st.container()
        
    with cont_dl_analyst:
        if os.path.exists(report_path):
            with open(report_path, "rb") as f:
                report_data = f.read()
            
            report_filename = os.path.basename(report_path)
            if st.download_button(
                label="Download Full Platform Report (Local Copy)",
                data=report_data,
                file_name=report_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="secondary"
            ):
                add_session_log(f"Downloaded {report_filename}")
                st.success("Report downloaded successfully.")
        else:
            st.info("Click 'Generate & Publish' above to create the first report.")

    render_footer("Export Reports")
