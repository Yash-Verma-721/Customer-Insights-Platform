from io import BytesIO
import pandas as pd

def _render_excel_sheet(worksheet, blocks):
    from openpyxl.styles import Font, PatternFill
    
    current_row = 1
    for b in blocks:
        if b["type"] == "title":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=20, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            worksheet.row_dimensions[current_row].height = 35
            current_row += 3
        elif b["type"] == "h1":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=16, bold=True, color="1E293B")
            current_row += 2
        elif b["type"] == "h2":
            cell = worksheet.cell(row=current_row, column=1, value=b["text"])
            cell.font = Font(size=12, bold=True, color="334155")
            current_row += 1
        elif b["type"] == "p":
            lines = str(b["text"]).split('\n')
            for line in lines:
                cell = worksheet.cell(row=current_row, column=1, value=line.strip())
                cell.font = Font(size=11, color="475569")
                current_row += 1
            current_row += 1
        elif b["type"] == "bullets":
            for item in b["items"]:
                cell = worksheet.cell(row=current_row, column=1, value=f"- {item}")
                cell.font = Font(size=11, color="475569")
                current_row += 1
            current_row += 1
        elif b["type"] == "table":
            cell = worksheet.cell(row=current_row, column=1, value=b["title"])
            cell.font = Font(bold=True, size=11, color="0F172A")
            current_row += 1
            
            df = b["df"]
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=str(col_name))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            current_row += 1
            
            for _, row in df.iterrows():
                for col_idx, val in enumerate(row, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=str(val))
                    cell.font = Font(size=11, color="334155")
                current_row += 1
            current_row += 2

def generate_excel_bytes(blocks, sheet_title="Report"):
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet(title=sheet_title)
        _render_excel_sheet(worksheet, blocks)
        
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
            
        worksheet.column_dimensions['A'].width = 90
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 30
        
    return excel_buffer.getvalue()

def build_module_report(title, metrics=None, insights=None, tables=None, charts=None, forecast=None, recommendations=None):
    """
    Construct standard report blocks used by every module.
    """
    blocks = []
    blocks.append({"type": "title", "text": title})
    
    section_num = 1
    
    if metrics:
        blocks.append({"type": "h1", "text": f"{section_num}. Executive KPIs"})
        metric_names = [m[0] for m in metrics]
        metric_values = [m[1] for m in metrics]
        blocks.append({"type": "table", "title": "Key Metrics", "df": pd.DataFrame({
            "Metric": metric_names,
            "Value": metric_values
        })})
        section_num += 1
        
    if tables and any(t.get("placement") == "before_insights" for t in tables):
        blocks.append({"type": "h1", "text": f"{section_num}. Summary Data"})
        for tbl in [t for t in tables if t.get("placement") == "before_insights"]:
            if tbl.get("df") is not None and not tbl.get("df").empty:
                blocks.append({"type": "table", "title": tbl.get("title", "Data"), "df": tbl["df"]})
            elif tbl.get("empty_message"):
                blocks.append({"type": "p", "text": tbl["empty_message"]})
        section_num += 1
        
    if insights:
        blocks.append({"type": "h1", "text": f"{section_num}. Insights"})
        if isinstance(insights, list):
            blocks.append({"type": "bullets", "items": insights})
        elif isinstance(insights, str):
            blocks.append({"type": "p", "text": insights})
        section_num += 1
        
    if forecast and forecast.get("status") == "success":
        blocks.append({"type": "h1", "text": f"{section_num}. AI Forecast"})
        
        forecast_df = pd.DataFrame({
            "Metric": ["Predicted Target", "Identified Trend", "Model Confidence", "Strategic Recommendation"],
            "Value": [
                forecast.get("formatted_target", forecast.get("forecast_revenue", "N/A")), 
                forecast.get("trend", "N/A"), 
                forecast.get("confidence", "N/A"), 
                forecast.get("recommendation", "N/A")
            ]
        })
        blocks.append({"type": "table", "title": "AI Prediction", "df": forecast_df})
        section_num += 1
        
    if tables and any(t.get("placement", "after_insights") == "after_insights" for t in tables):
        blocks.append({"type": "h1", "text": f"{section_num}. Detailed Data"})
        for tbl in [t for t in tables if t.get("placement", "after_insights") == "after_insights"]:
            if tbl.get("df") is not None and not tbl.get("df").empty:
                blocks.append({"type": "table", "title": tbl.get("title", "Data"), "df": tbl["df"]})
            elif tbl.get("empty_message"):
                blocks.append({"type": "p", "text": tbl["empty_message"]})
        section_num += 1
        
    if recommendations:
        blocks.append({"type": "h1", "text": f"{section_num}. Recommendations"})
        blocks.append({"type": "bullets", "items": recommendations})
        section_num += 1
        
    return blocks
